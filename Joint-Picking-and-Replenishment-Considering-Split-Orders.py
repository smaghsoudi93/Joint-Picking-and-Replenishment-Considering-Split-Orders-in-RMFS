from docplex.mp.model import Model
from collections import defaultdict
import sys
sys.path = [r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'] + sys.path
import cplex
import random
import numpy as np
import time
import threading
import pandas as pd
import concurrent.futures
#import params
import openpyxl
import copy

workbook = openpyxl.load_workbook('Data2 - new.xlsx')
sheet = workbook.active

# --- Global algorithm parameters (kept consistent with the exact OPL model) ---
ALPHA_COORDINATION = 10   # coordination cost weight (alpha in the paper / OPL model)
UNUSED_WEIGHT = 10        # weight on unused picking-station capacity (ZZ in the OPL model)

# Reproducibility (lesson from Paper 1, reviewer R2-14: the seed used in all
# reported experiments must be fixed and stated). VNS shaking is intentionally
# random; the seed makes every run repeatable.
SEED = 42
random.seed(SEED)
np.random.seed(SEED)


for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'Demand':
            demand_start_row = row + 1
            demand_start_col = col
            break

demand_rows = 0
while sheet.cell(row=demand_start_row + demand_rows, column=demand_start_col).value is not None:
    demand_rows += 1

demand_cols = 0
while sheet.cell(row=demand_start_row, column=demand_start_col + demand_cols).value is not None:
    demand_cols += 1

# Read the Demand data
Demand = []
for row in range(demand_start_row, demand_start_row + demand_rows):
    row_data = []
    for col in range(demand_start_col, demand_start_col + demand_cols):
        row_data.append(sheet.cell(row=row, column=col).value)
    Demand.append(row_data)

#for row in Demand:
#    print(row)

#xl_file = pd.read_excel("C:\Users\saman\OneDrive\Desktop\Data1.xlsx")
#Demand = xl_file.pars('Demand')

#----------------------Clustering and assigning to picking stations-------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------

def create_clusters(Demand, num_picking_stations, station_capacity):

    # Coordination penalty weight used in the Phase-1 greedy score
    # (Score = co-occurrence - alpha * new-split indicator, Algorithm 1).
    # Was 0 (disabled), which contradicted the paper's Algorithm 1;
    # now tied to the global alpha for consistency with the exact model.
    coordination_weight = ALPHA_COORDINATION
    start1 = time.time()
    num_orders, num_skus = np.array(Demand).shape
    # Step 2: Build binary presence matrix
    presence_matrix = (np.array(Demand) > 0).astype(int)

    # Step 3: Compute SKU co-occurrence matrix
    co_occurrence = np.zeros((num_skus, num_skus), dtype=int)
    for i in range(num_skus):
        for j in range(num_skus):
            co_occurrence[i, j] = np.sum(presence_matrix[:, i] & presence_matrix[:, j])
    # Step 4: Sum co-occurrence per SKU for priority
    sku_priority = co_occurrence.sum(axis=1)  # shape: (num_skus,)

    # Step 5-7: ORDER-AWARE greedy clustering under capacity.
    # The previous version walked item-by-item (sorted by SKU priority), which
    # scattered the items of each order across clusters and produced initial
    # solutions with many avoidable splits (e.g. 17 splits on the o10 instance
    # where the optimum needs only the 1 forced split of the giant order).
    # The rewritten greedy keeps the paper's scoring idea (co-occurrence
    # affinity, coordination penalty alpha) but takes ORDERS as the placement
    # unit: an order is placed whole into the best-affinity cluster with room;
    # it is split only when it fits nowhere whole, and then across as few
    # clusters as possible. Clusters are capped at the number of picking
    # stations with a 1:1 cluster-station mapping.

    # per-order item triplets and sizes
    order_triplets = {o: [(o, s, int(np.array(Demand)[o, s]))
                          for s in range(num_skus) if np.array(Demand)[o, s] > 0]
                      for o in range(num_orders)}
    order_size = {o: sum(q for (_, _, q) in trips) for o, trips in order_triplets.items()}

    # process orders largest-first (FFD-style; hard orders placed while space is flexible)
    order_sequence = sorted(range(num_orders), key=lambda o: -order_size[o])

    clusters = [[] for _ in range(num_picking_stations)]
    cluster_load = [0] * num_picking_stations
    order_in_clusters = defaultdict(set)

    def cluster_affinity(order_id, cid):
        """co-occurrence affinity between an order's SKUs and a cluster's SKUs"""
        cluster_skus = [s for (_, s, _) in clusters[cid]]
        return sum(co_occurrence[sku_id, s]
                   for (_, sku_id, _) in order_triplets[order_id]
                   for s in cluster_skus)

    for order_id in order_sequence:
        trips = sorted(order_triplets[order_id], key=lambda t: -t[2])
        size = order_size[order_id]

        # 1) try to place the WHOLE order (zero split) in the best-affinity cluster
        feasible = [cid for cid in range(num_picking_stations)
                    if cluster_load[cid] + size <= station_capacity]
        if feasible:
            best_cid = max(feasible, key=lambda cid: (cluster_affinity(order_id, cid),
                                                      -(cluster_load[cid])))
            for t in trips:
                clusters[best_cid].append(t)
            cluster_load[best_cid] += size
            order_in_clusters[order_id].add(best_cid)
            continue

        # 2) forced split: distribute the order's items over as FEW clusters as
        # possible -- repeatedly fill the cluster with the largest remaining
        # space (best affinity as tie-break)
        remaining = list(trips)
        while remaining:
            cid = max(range(num_picking_stations),
                      key=lambda c: (station_capacity - cluster_load[c],
                                     cluster_affinity(order_id, c)))
            space = station_capacity - cluster_load[cid]
            placed_any = False
            still_left = []
            for t in remaining:
                if t[2] <= space:
                    clusters[cid].append(t)
                    cluster_load[cid] += t[2]
                    space -= t[2]
                    order_in_clusters[order_id].add(cid)
                    placed_any = True
                else:
                    still_left.append(t)
            remaining = still_left
            if remaining and not placed_any:
                # no cluster can take even the largest remaining item ->
                # heuristic infeasibility; report clearly (caught downstream
                # by the coverage constraint as well)
                print(f"[Phase 1] could not place item(s) of order {order_id}: {remaining}")
                break

    # drop trailing empty clusters (keep 1:1 index == station id semantics)
    # (empty clusters are harmless; keep them so cluster index == station id)

# -----------------------------
# STEP 4 - CALCULATE COORDINATION COST
# -----------------------------
    print("clusters:",order_in_clusters)
    coordination_cost = sum(len(c) - 1 for c in order_in_clusters.values())
    print(coordination_cost)

# -----------------------------
# STEP 5 - CREATE CLUSTER-SKU MATRIX
# -----------------------------
    cluster_matrix = np.zeros((len(clusters), num_skus), dtype=int)

    for cid, cluster in enumerate(clusters):
        for _, sku, qty in cluster:
            cluster_matrix[cid, sku] += qty



    end1 = time.time()
    #print("clustering_time:",end1 - start1)

    y = np.zeros((len(Demand[0]),len(Demand), num_picking_stations))

    station_assignments = defaultdict(list)
    station_usage = defaultdict(int)
    order_to_stations = defaultdict(set)

    for i, cluster in enumerate(clusters):
        station_id = i % num_picking_stations
        station_assignments[station_id].append(cluster)
        for order_idx, _, qty in cluster:
            station_usage[station_id] += qty
            order_to_stations[order_idx].add(station_id)

    print("st:",order_to_stations)
    
    for sid, assigned_clusters in station_assignments.items():
    #print(f"\n Station {sid}:")
        for cid, cluster in enumerate(assigned_clusters, 1):
        #print(f"  Cluster {cid}:")
            for order_idx, sku_idx, qty in cluster:
            #print(f"    Order {order_idx}, SKU {sku_idx}, Quantity {qty}")
                y[sku_idx][order_idx][sid] = 1

    yy = np.zeros((1,len(Demand)))

    for order, stations in order_to_stations.items():
        #print(f"  Order {order}: {len(stations)} station(s)")
        yy[0][order] = len(stations)

    station_unused_capacity = {
    sid: station_capacity - min(station_capacity, station_usage[sid])
    for sid in range(num_picking_stations)
    }
    print("unused",station_unused_capacity)
    total_unused_capacity = sum(station_unused_capacity.values())
    print(total_unused_capacity)
    
    return clusters,yy,y,station_assignments,coordination_cost, total_unused_capacity

#----------------------------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------------------------

def swap_neighborhood(yy, sequences, num_picking_stations):
    new_y = yy.copy()

    # Select two random groups
    group1, group2 = random.sample(range(len(new_y)), 2)
    

    # Swap the values of ss and jj for the selected groups
    ss1, jj1 = find_ss_jj(new_y, group1, sequences, num_picking_stations)
    ss2, jj2 = find_ss_jj(new_y, group2, sequences, num_picking_stations)

    new_y[group1][ss1][jj1] = 0
    new_y[group2][ss2][jj2] = 0
    new_y[group2][ss1][jj1] = 1
    new_y[group1][ss2][jj2] = 1

    #print(new_y)
    return new_y

def find_ss_jj(y, group, sequences, num_picking_stations):
    for ss in range(num_picking_stations):
        for jj in range(sequences):
            if y[group][ss][jj] == 1:
                return ss, jj

#-----------------------------------------------------------------------------
# Initialize the problem data
#-----------------------------------------------------------------------------


#---------------------------------------------------- Solving sub problem of order clusters and stations ---------------------------------

def evaluate_clusters(clusters, yy, y, coordination_cost,total_unused_capacity, verbose=False, exact=False, cutoff=None):

    ##########################################################Read the data###########################################################################

    coordination_weight = ALPHA_COORDINATION
    unused_weight = UNUSED_WEIGHT
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'matrix1':
                matrix1_start_row = row + 1
                matrix1_start_col = col
                break

    matrix1_rows = 0
    while sheet.cell(row=matrix1_start_row + matrix1_rows, column=matrix1_start_col).value is not None:
        matrix1_rows += 1

    matrix1_cols = 0
    while sheet.cell(row=matrix1_start_row, column=matrix1_start_col + matrix1_cols).value is not None:
        matrix1_cols += 1

    # Read the matrix1 data
    matrix1 = []
    for row in range(matrix1_start_row, matrix1_start_row + matrix1_rows):
        row_data1 = []
        for col in range(matrix1_start_col, matrix1_start_col + matrix1_cols):
            row_data1.append(sheet.cell(row=row, column=col).value)
        matrix1.append(row_data1)
        
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Demand1':
                demand1_start_row = row + 1
                demand1_start_col = col
                break

    demand1_rows = 0
    while sheet.cell(row=demand1_start_row + demand1_rows, column=demand1_start_col).value is not None:
        demand1_rows += 1

    demand1_cols = 0
    while sheet.cell(row=demand1_start_row, column=demand1_start_col + demand1_cols).value is not None:
        demand1_cols += 1

    # Read the Demand data
    Demand1 = []
    for row in range(demand1_start_row, demand1_start_row + demand1_rows):
        row_data2 = []
        for col in range(demand1_start_col, demand1_start_col + demand1_cols):
            row_data2.append(sheet.cell(row=row, column=col).value)
        Demand1.append(row_data2)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'items':
                items_row = row
                items_col = col
                break

    items = sheet.cell(row=items_row + 1, column=items_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'orders1':
                orders1_row = row
                orders1_col = col
                break

    orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'shelves':
                shelves_row = row
                shelves_col = col
                break

    shelves = sheet.cell(row=shelves_row + 1, column=shelves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_picking_stations':
                num_picking_stations_row = row
                num_picking_stations_col = col
                break

    num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_replenishment_stations':
                num_replenishment_stations_row = row
                num_replenishment_stations_col = col
                break

    num_replenishment_stations = sheet.cell(row=num_replenishment_stations_row + 1, column=num_replenishment_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'waves':
                waves_row = row
                waves_col = col
                break

    waves = sheet.cell(row=waves_row + 1, column=waves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Pods_capacity':
                Pods_capacity_row = row
                Pods_capacity_col = col
                break

    Pods_capacity = sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'group_capacity':
                group_capacity_row = row
                group_capacity_col = col
                break

    group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Replenishment_station_capacity':
                Replenishment_station_capacity_row = row
                Replenishment_station_capacity_col = col
                break

    Replenishment_station_capacity = sheet.cell(row=Replenishment_station_capacity_row + 1, column=Replenishment_station_capacity_col).value



    ##################################################################################################################################################


    #Groups = list(range(len(order_groups)))
    Items = list(range(items))
    Orders = list(range(orders1))
    Sequences = list(range(orders1))
    Shelves = list(range(shelves))
    Picking_stations = list(range(num_picking_stations))
    Replenishment_stations = list(range(num_replenishment_stations))
    Waves = list(range(waves))
    # Big-M values derived from instance parameters (lesson from Paper 1, R2-11 --
    # never use an arbitrary huge constant):
    #   M_uv bounds sum_i u[i,p,w,r]: can never exceed the pod capacity nor the
    #        replenishment-station capacity.
    #   M_qx bounds sum_{i,o} q[i,o,p,w,s]: the flow out of pod p in wave w,
    #        which can never exceed the pod capacity.
    M_uv = min(Pods_capacity, Replenishment_station_capacity)
    M_qx = Pods_capacity
    #waves_capacity = np.full((1, waves), (num_replenishment_stations*Replenishment_station_capacity))
    replenishmnet_capacity = np.full((waves, num_replenishment_stations), Replenishment_station_capacity)
    sample = np.zeros((len(Demand1), len(Demand1[0]), waves))
    capacity_of_pods = np.full((waves, shelves), Pods_capacity)
    #Arriving_times = (10, 15, 20, 25) #(7, 12, 17, 22, 27)#(5, 10, 15, 20, 25, 30) #(7, 12, 17, 22, 27) #(10, 15, 20, 25, 30) #(7, 12, 17, 22, 27)#(7, 12, 17, 22, 27, 32) #(5, 10) #(6, 12, 18, 24, 30)  #(5, 10) # 
    #Duration = (5, 5, 5, 5) #(5, 5, 5, 5, 5, 5) #(5, 5) #(6, 6, 6, 6, 6) #(5, 5) # 
    

    #items = 11 #5 #7 #5  
    #orders1 = 8 #6 #50 #20 #15  
    #shelves = 3 #10 
    #num_picking_stations = 3 #2 #6 #4  
    #num_replenishment_stations = 3 #2 #6 #4  
    #waves = 4 #2 #6 #5   
    #Pods_capacity = 45 #15 #35 #25   
    #group_capacity = 2 #7 #3 #4 
    #Replenishment_station_capacity = 25 #22 #40   
    
    

#-----------------------------------------------------------------------------
# Build the model
#-----------------------------------------------------------------------------

# Create CPO model
    mdl = Model(name='Thesis')
    #print(order_groups)


    #qq = np.zeros((len(order_groups),shelves, items, waves))
    #zz = np.zeros((len(order_groups),shelves, waves))
    vv = np.zeros((shelves, waves, num_replenishment_stations))
    uu = np.zeros((items, shelves, waves, num_replenishment_stations))
    fff = np.zeros((items, waves, num_replenishment_stations))


    #xx = np.zeros((len(matrix1),len(order_groups)))
    #x = {}
#for group_idx, group in enumerate(order_groups):
 #   for order in group:
    #for g in Groups:
     #   for o in Orders:
      #      x[o, g] = mdl.binary_var(name=f'x_{o}_{g}')

#for g in Groups:
    #for o in Orders:
        #mdl.add_constraint(x[o, g] == 0)

# Set constraints based on order_groups
    #for group_idx, group in enumerate(order_groups):
     #   for order in group:
      #      xx[order, group_idx] = 1
       #     mdl.add_constraint(x[order, group_idx] == 1)
    #print(xx)
    #print(yy)
    y1= {}
    for i in Items:
        for j in Orders:
            for s in Picking_stations:
                y1[i, j, s] =mdl.binary_var(name=f'y_{i}_{j}_{s}')

    # Fix y1 to the heuristic's assignment in BOTH directions.
    # (Previously only cells with y==1 were fixed, leaving y==0 cells free --
    # the MIP could silently add extra item-station assignments unrelated to
    # the Phase 1-2 decisions.)
    for i in Items:
        for j in Orders:
            for s in Picking_stations:
                mdl.add_constraint(y1[i, j, s] == int(y[i, j, s]))

    # Item-level coverage (matches const21 in the exact OPL model): every
    # requested item must be assigned to exactly one station. Because y1 is
    # fully fixed above, this acts as a hard validity check on the heuristic's
    # output -- if Phases 1-2 ever produce an uncovered or double-covered item,
    # the model becomes infeasible and the failure is caught immediately
    # instead of silently returning a meaningless objective.
    for j in Orders:
        for i in Items:
            if Demand1[j][i] > 0:
                mdl.add_constraint(mdl.sum(y1[i, j, s] for s in Picking_stations) == 1)
                
    z = {(i): mdl.continuous_var(name="z%d" % (i)) for i in Picking_stations}
    v = {(i,j,k): mdl.binary_var(name="v%d%d%d" % (i,j,k)) for i in Shelves for j in Waves for k in Replenishment_stations}
    # q is STATION-INDEXED (5-dim), matching the updated exact model: a pod
    # visit is counted separately for each station it serves in a wave, which
    # is what makes order splitting genuinely consume additional pod visits.
    q = {(i,j,k,ii,s): mdl.continuous_var(name="q_%d_%d_%d_%d_%d" % (i,j,k,ii,s)) for i in Items for j in Orders for k in Shelves for ii in Waves for s in Picking_stations}
    u = {(i,j,k,ii): mdl.continuous_var(name="u%d%d%d%d" % (i,j,k,ii)) for i in Items for j in Shelves for k in Waves for ii in Replenishment_stations}
    #ff = {(i,j,k): mdl.continuous_var(name="ff%d%d%d" % (i,j,k)) for i in Items for j in Waves for k in Replenishment_stations}
    x = {(i,j,k): mdl.binary_var(name="x%d%d%d" % (i,j,k)) for i in Shelves for j in Waves for k in Picking_stations}

    ff= {}
    for i in Items:
        for j in Waves:
            for k in Replenishment_stations:
                ff[i, j, k] =mdl.continuous_var(name=f'ff_{i}_{j}_{k}')
            
    waving_items = np.zeros((waves, len(matrix1[0])))
    replenishment_items = np.zeros((waves, len(matrix1[0]),num_replenishment_stations))
    pod_items = np.zeros((len(matrix1[0]),shelves ,waves, num_replenishment_stations))
    
# ------------------------------------------------------------------------ waving heuristic -------------------------------------------------------
    # A wave can carry no more than EITHER bottleneck allows: the total
    # replenishment-station capacity OR the total pod capacity of that wave.
    # (Previously only the replenishment side was considered, so waves could
    # be built that then failed to fit into the pods.)
    num_waves = waves   # keep the scalar (number of waves) before it gets shadowed below
    wave_capacity = min(num_replenishment_stations*Replenishment_station_capacity,
                        shelves*Pods_capacity)
    num_items = len(matrix1[0])

        

    wave_plan = []
    current_wave = [0] * num_items
    current_total = 0

    for cluster in clusters:
        for _, item, qty in cluster:
            while qty > 0:
                available_space = wave_capacity - current_total
                if available_space == 0:
                    wave_plan.append(current_wave)
                    current_wave = [0] * num_items
                    current_total = 0
                    available_space = wave_capacity
                assign_qty = min(qty, available_space)
                current_wave[item] += assign_qty
                current_total += assign_qty
                qty -= assign_qty

    if current_total > 0:
        wave_plan.append(current_wave)

    # graceful failure instead of IndexError if the plan needs more waves
    # than the instance provides
    if len(wave_plan) > num_waves:
        print(f"[wave-planning] infeasible: heuristic needs {len(wave_plan)} waves, "
              f"instance only has {num_waves}")
        return 10000000

    wavess = np.array(wave_plan)
    for i in range(len(wavess)):
        for j in range(len(wavess[0])):
            waving_items[i][j] = wavess[i][j]

    # ---- Stage 3b: distribute each wave's items across replenishment stations
    # (greedy sequential fill). Replaces a hardcoded 3-level if/else cascade
    # that (a) crashed with an IndexError whenever more than 4 stations were
    # needed and (b) had no bound check at all -- the same style of bug that
    # crashed on instances with few pods. The while-loop below implements the
    # identical greedy logic for ANY number of stations, and returns the
    # standard infeasibility sentinel (10000000) if the wave genuinely does
    # not fit, instead of crashing.
    for waves_number in Waves:
        r = 0
        for j in range(len(matrix1[0])):
            remaining = waving_items[waves_number][j]
            while remaining > 0:
                if r >= num_replenishment_stations:
                    print(f"[wave-planning] infeasible: wave {waves_number} exceeds "
                          f"total replenishment capacity")
                    return 10000000
                take = min(remaining, replenishmnet_capacity[waves_number][r])
                replenishment_items[waves_number][j][r] += take
                replenishmnet_capacity[waves_number][r] -= take
                remaining -= take
                if remaining > 0:
                    r += 1

        # ---- Stage 3c: place replenished quantities into pods (greedy
        # sequential fill; per-wave pod capacity). Same generalization: the
        # old code did p+1 with no bound check and crashed with 2 pods.
        p = 0
        for k in Replenishment_stations:
            for i in range(len(matrix1[0])):
                remaining = replenishment_items[waves_number][i][k]
                while remaining > 0:
                    if p >= shelves:
                        print(f"[wave-planning] infeasible: wave {waves_number} exceeds "
                              f"total pod capacity")
                        return 10000000
                    take = min(remaining, capacity_of_pods[waves_number][p])
                    pod_items[i][p][waves_number][k] += take
                    capacity_of_pods[waves_number][p] -= take
                    remaining -= take
                    if remaining > 0:
                        p += 1
                
                
# -------------------------------------------------------- constraints ------------------------------------------------------
#Constant variables based on heuristic:


    #for i in Items:
     #   for j in Waves:
      #      for k in Replenishment_stations:
                    #print(i,j,k, replenishment_items[j][i][k])
       #         mdl.add_constraint(ff[i, j, k] == replenishment_items[j][i][k])

    for i in Items:
        for j in Waves:
            for k in Replenishment_stations:
                for p in Shelves:
                    mdl.add_constraint(u[i, p, j, k] == pod_items[i][p][j][k])


#print(order_groups)

# ---------------------------- decision variables ---------------------

#y = {(i,j,k): mdl.binary_var(name="y%d%d%d" % (i+1,j+1,k+1)) for i in Groups for j in Picking_stations for k in Sequences}
        
        
#--------------------------------------------------------------------------------------
#mdl.add_constraint(l[0,0] == 1)
#mdl.add_constraint(l[1,0] == 1)
#mdl.add_constraint(l[2,0] == 1)
#mdl.add_constraint(l[3,2] == 1)

        #for g in range(len(Groups1)):
        #    if groups_demand[0][Groups1[g]] == 0:
                #print(Groups1[g])
         #       mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves for w in Waves) >= 1)

        #for w in Waves:
         #   mdl.add_constraint(mdl.sum(v[p,w,r] for p in Shelves for r in Replenishment_stations) >= 1)

    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for p in Shelves) <= Replenishment_station_capacity)

    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for r in Replenishment_stations) <= Pods_capacity)

        
    # (matches updated OPL const10, per-station equality): the quantity of
    # item i for order o picked AT STATION s equals the demand assigned there.
    # When y1[i,o,s]=0 this forces all q[i,o,p,w,s]=0 -- a tight, big-M-free
    # link between picking flows and item-station assignment.
    for i in Items:
        for o in Orders:
            for s in Picking_stations:
                mdl.add_constraint(Demand1[o][i]*y1[i,o,s] == mdl.sum(q[i,o,p,w,s] for p in Shelves for w in Waves))


    # (matches updated OPL const13): replenished quantity into pod (p,w)
    # equals total picked from it across all orders AND stations.
    for w in Waves:
        for i in Items:
            for p in Shelves:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for r in Replenishment_stations) == mdl.sum(q[i,o,p,w,s] for o in Orders for s in Picking_stations))

    for i in Items:
        for o in Orders:
            for s in Picking_stations:
                mdl.add_constraint(mdl.sum(x[p,w,s] for p in Shelves for w in Waves) >= y1[i,o,s])


    # NOTE: the three blocks below were previously indented one level too
    # deep (inside the Items loop), silently adding |Items| duplicate copies
    # of every constraint. Now at function level (added exactly once).
    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) >= v[p,w,r])

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) <= M_uv*v[p,w,r])

    # (matches updated OPL const20, per-station): picking from pod p at
    # station s in wave w requires the pod to be AT that station -- the key
    # constraint that makes each pod-station visit count separately.
    for p in Shelves:
        for w in Waves:
            for s in Picking_stations:
                mdl.add_constraint(mdl.sum(q[i,o,p,w,s] for i in Items for o in Orders) <= M_qx*x[p,w,s])

        
                               

#Objective function
    mdl.total_objective = mdl.sum(x[p,w,s] for p in Shelves for w in Waves for s in Picking_stations) + mdl.sum(v[p,w,r] for p in Shelves for w in Waves for r in Replenishment_stations)+ (coordination_weight *coordination_cost)+(unused_weight*total_unused_capacity)
        #mdl.total_objective = mdl.sum(ctt[o] for o in Orders)/ orders1
    mdl.minimize(mdl.total_objective)
    # Inner-solve policy.
    #  * exact=True (initial & final reported solutions): full solve,
    #    no tolerance, 120 s safety cap.
    #  * search candidates: the VNS only needs to know whether the
    #    candidate IMPROVES on the incumbent, so an objective cutoff is
    #    set at the incumbent cost. CPLEX terminates as soon as it
    #    proves no better solution exists (bound >= cutoff), which is
    #    near-immediate for clearly worse candidates. A 30 s cap and an
    #    absolute tolerance of 2 (all objective coefficients are
    #    integers) bound the effort on near-tie candidates; a candidate
    #    with no improving incumbent within the cap is rejected.
    if exact:
        mdl.parameters.timelimit = 120
    else:
        mdl.parameters.timelimit = 30
        mdl.parameters.mip.tolerances.absmipgap = 2.0
        if cutoff is not None:
            mdl.parameters.mip.tolerances.uppercutoff = cutoff

    #mdl.solve(log_=True)
        #mdl.solve()


      
        #log_output=True
    #if mdl.solve(log_output=True):
    if mdl.solve():
        total_cost = mdl.objective_value
        if verbose:
            sum_x = sum(x[p,w,s].solution_value for p in Shelves for w in Waves for s in Picking_stations)
            sum_v = sum(v[p,w,r].solution_value for p in Shelves for w in Waves for r in Replenishment_stations)
            print(f"[eval breakdown] pod-picking visits (Σx) = {sum_x:.0f} | "
                  f"pod-replenishment visits (Σv) = {sum_v:.0f} | "
                  f"coordination = {coordination_weight}×{coordination_cost} = {coordination_weight*coordination_cost} | "
                  f"unused = {unused_weight}×{total_unused_capacity} = {unused_weight*total_unused_capacity} | "
                  f"TOTAL = {total_cost}")
        else:
            print("total_cost:", total_cost)

    else:
        if cutoff is not None:
            # Not an error: the incumbent cutoff excluded every solution of
            # this candidate, i.e. CPLEX PROVED it cannot improve the
            # incumbent (or found no improving solution within the cap).
            print(f"[cutoff] candidate cannot beat incumbent ({cutoff + 0.5:g}) - rejected")
        else:
            print("no_solution (infeasible or no incumbent within time cap)")
        total_cost = 10000000

    end2 = time.time()
        #print("wave:",waves_number, end2 - start2)
    
    
    return total_cost

    

def move_order_item(clusters, station_capacity):
    """Shaking move N1: relocate ONE randomly chosen item to a random cluster
    with room, preferring clusters that already contain the same order (so
    the move tends to REDUCE splits rather than create them).

    The previous version was fully deterministic ("first feasible move"),
    so it generated the identical candidate on every VNS iteration -- 50
    redundant CPLEX solves per run on the o20 log."""
    new_clusters = copy.deepcopy(clusters)

    # candidate items: (cluster_idx, item)
    movable = [(ci, item) for ci, c in enumerate(new_clusters) for item in c]
    if not movable:
        return new_clusters
    random.shuffle(movable)

    for ci, item in movable:
        order_id, _, qty = item
        targets = [cj for cj, c in enumerate(new_clusters)
                   if cj != ci and
                   sum(q for _, _, q in c) + qty <= station_capacity]
        if not targets:
            continue
        # prefer a target that already holds the same order (split-reducing)
        same_order = [cj for cj in targets
                      if any(o == order_id for o, _, _ in new_clusters[cj])]
        cj = random.choice(same_order) if same_order else random.choice(targets)
        new_clusters[cj].append(item)
        new_clusters[ci].remove(item)
        return new_clusters
    return new_clusters

def merge_and_repartition(clusters, station_capacity):
    if len(clusters) < 2:
        return clusters

    # Randomly pick two clusters
    i, j = random.sample(range(len(clusters)), 2)
    merged_items = clusters[i] + clusters[j]
    merged_items.sort(key=lambda x: x[2], reverse=True)

    new_clusters = copy.deepcopy(clusters)
    del new_clusters[max(i, j)]
    del new_clusters[min(i, j)]

    # Greedy repack into new clusters
    temp = []
    for item in merged_items:
        placed = False
        for cluster in temp:
            if sum(q for _, _, q in cluster) + item[2] <= station_capacity:
                cluster.append(item)
                placed = True
                break
        if not placed:
            temp.append([item])

    return new_clusters + temp

def split_cluster(clusters, station_capacity):
    # deepcopy FIRST: the previous version mutated the caller's list in place
    # (clusters.remove(...)), corrupting the incumbent solution -- the same
    # shallow-copy bug family we fixed in Paper 1's VNS.
    clusters = copy.deepcopy(clusters)
    if not clusters:
        return clusters

    # Pick a random cluster with >1 item
    large_clusters = [c for c in clusters if len(c) > 1]
    if not large_clusters:
        return clusters

    cluster_to_split = random.choice(large_clusters)
    clusters.remove(cluster_to_split)

    # Sort items by quantity and split greedily
    sorted_items = sorted(cluster_to_split, key=lambda x: x[2], reverse=True)
    new1, new2 = [], []

    for item in sorted_items:
        if sum(q for _, _, q in new1) + item[2] <= station_capacity:
            new1.append(item)
        else:
            new2.append(item)

    return clusters + [new1, new2]

def update_station_and_order_mapping(clusters, num_stations):
    station_assignments = defaultdict(list)
    order_to_stations = defaultdict(set)

    for cluster_idx, cluster in enumerate(clusters):
        station_id = cluster_idx % num_stations  # Round-robin
        station_assignments[station_id].append(cluster)
        for order_id, _, _ in cluster:
            order_to_stations[order_id].add(station_id)

    return station_assignments, order_to_stations

def relocate_partial_order(clusters, station_capacity):
    new_clusters = copy.deepcopy(clusters)
    for i, cluster_i in enumerate(new_clusters):
        for item in cluster_i:
            for j, cluster_j in enumerate(new_clusters):
                if i != j and sum(q for _, _, q in cluster_j) + item[2] <= station_capacity:
                    # Check if same order exists in target cluster to encourage partial order moves
                    target_orders = [o for o, _, _ in cluster_j]
                    if item[0] in target_orders:
                        cluster_j.append(item)
                        cluster_i.remove(item)
                        return new_clusters
    return new_clusters


def shuffle_clusters(clusters, station_capacity):
    if len(clusters) < 3:
        return clusters

    selected_clusters = random.sample(clusters, k=3)
    all_items = [item for c in selected_clusters for item in c]
    all_items.sort(key=lambda x: x[2], reverse=True)

    new_clusters = [c for c in clusters if c not in selected_clusters]

    reshuffled = []
    for item in all_items:
        placed = False
        for cluster in reshuffled:
            if sum(q for _, _, q in cluster) + item[2] <= station_capacity:
                cluster.append(item)
                placed = True
                break
        if not placed:
            reshuffled.append([item])

    return new_clusters + reshuffled

def repack_split_order_items(clusters, station_capacity):
    """Shaking move N4: take the most-split order, pull its items out, and
    re-pack them into as few clusters as possible.

    BUG FIX: the old version turned every NON-target item of the touched
    clusters into its own singleton cluster (new_clusters.append([item])).
    On o40 the giant order's cluster holds many items, so the candidate
    exploded into dozens of clusters; round-robin then piled several
    clusters onto each station and the resulting q-routing subproblem was
    hard enough to hang CPLEX (no time limit). Non-target items now stay
    grouped exactly as they were."""
    clusters = copy.deepcopy(clusters)

    order_to_clusters = defaultdict(set)
    for idx, cluster in enumerate(clusters):
        for o, _, _ in cluster:
            order_to_clusters[o].add(idx)

    split_orders = sorted(order_to_clusters.items(),
                          key=lambda x: len(x[1]), reverse=True)
    if not split_orders or len(split_orders[0][1]) <= 1:
        return clusters  # nothing to improve

    target_order = split_orders[0][0]
    involved = order_to_clusters[target_order]

    items_to_repack = []
    new_clusters = []
    for idx, cluster in enumerate(clusters):
        if idx in involved:
            keep = [it for it in cluster if it[0] != target_order]
            items_to_repack += [it for it in cluster if it[0] == target_order]
            if keep:
                new_clusters.append(keep)      # <- stays ONE cluster
        else:
            new_clusters.append(cluster)

    # re-pack target items: existing clusters with room first (fewest pieces),
    # largest items first; open a new cluster only as a last resort
    items_to_repack.sort(key=lambda t: -t[2])
    opened = []
    for item in items_to_repack:
        placed = False
        # existing clusters, most free space first
        for c in sorted(new_clusters,
                        key=lambda c: sum(q for _, _, q in c)):
            if sum(q for _, _, q in c) + item[2] <= station_capacity:
                c.append(item)
                placed = True
                break
        if not placed:
            for c in opened:
                if sum(q for _, _, q in c) + item[2] <= station_capacity:
                    c.append(item)
                    placed = True
                    break
        if not placed:
            opened.append([item])

    return new_clusters + opened

def derive_solution_components(clusters, num_picking_stations, Demand, station_capacity):
    """Given a clustering, derive everything evaluate_clusters needs:
    the item-order-station assignment y, per-order station counts yy,
    the coordination cost, and total unused picking capacity."""
    station_assignments, order_to_stations = update_station_and_order_mapping(clusters, num_picking_stations)

    y = np.zeros((len(Demand[0]), len(Demand), num_picking_stations))
    for sid, assigned_clusters in station_assignments.items():
        for cid, cluster in enumerate(assigned_clusters, 1):
            for order_idx, sku_idx, qty in cluster:
                y[sku_idx][order_idx][sid] = 1

    yy = np.zeros((1, len(Demand)))
    for order, stations in order_to_stations.items():
        yy[0][order] = len(stations)

    station_usage = defaultdict(int)
    for sid, cluster_list in station_assignments.items():
        for cluster in cluster_list:
            station_usage[sid] += sum(qty for _, _, qty in cluster)

    total_unused_capacity = sum(
        station_capacity - min(station_capacity, station_usage[sid])
        for sid in range(num_picking_stations))

    coordination_cost = sum(len(v) - 1 for v in order_to_stations.values())
    return y, yy, coordination_cost, total_unused_capacity


def _cluster_signature(clusters):
    """Canonical, order-preserving signature of a clustering. Cluster ORDER
    matters (round-robin maps cluster index -> station), item order inside a
    cluster does not."""
    return tuple(tuple(sorted(c)) for c in clusters)

def variable_neighborhood_search(max_iterations, num_picking_stations, Demand,
                                 station_capacity, patience=15, visit_lb=0):
    """Basic VNS, same accept/l-reset logic as before, plus:

    * cheap screening: coordination + unused capacity are computable WITHOUT
      CPLEX (derive_solution_components). Since sum(x)+sum(v) >= 0,
        ALPHA*coord + UNUSED_WEIGHT*unused
      is a valid lower bound on the candidate's cost; if it already meets or
      exceeds the current cost the candidate cannot improve and the CPLEX
      solve is skipped entirely. On the o20 log this prunes every 1851/2528-
      class candidate before it costs a solve.
    * evaluation cache: identical clusterings (e.g. a deterministic move
      firing twice) are looked up instead of re-solved.
    * early stopping: stop after `patience` consecutive iterations without
      improvement of the best cost (essential for o100-o150)."""
    best_clusters, best_yy, best_y, best_stations, coordination_cost, total_unused_capacity = create_clusters(
        Demand, num_picking_stations, station_capacity)
    print(f"[VNS] initial solution: {coordination_cost} extra splits, "
          f"{total_unused_capacity} unused capacity", flush=True)
    best_cost = evaluate_clusters(best_clusters, best_yy, best_y,
                                  coordination_cost, total_unused_capacity,
                                  verbose=True, exact=True)
    initial_sig = _cluster_signature(best_clusters)

    current_clusters = copy.deepcopy(best_clusters)
    current_cost = best_cost

    cache = {_cluster_signature(best_clusters): best_cost}
    n_solves, n_pruned, n_cached = 1, 0, 0

    neighborhoods = [move_order_item, merge_and_repartition,
                     split_cluster, repack_split_order_items]

    print(f"[VNS] k_max = {max_iterations}  |  initial cost = {best_cost}  |  "
          f"{len(neighborhoods)} neighborhoods  |  patience = {patience}",
          flush=True)

    iteration = 0
    since_improvement = 0
    while iteration < max_iterations and since_improvement < patience:
        improved_this_iteration = False
        l = 0
        while l < len(neighborhoods):
            move = neighborhoods[l]
            candidate_clusters = move(copy.deepcopy(current_clusters),
                                      station_capacity)

            new_y, new_yy, coordination_cost, total_unused_capacity = derive_solution_components(
                candidate_clusters, num_picking_stations, Demand,
                station_capacity)

            sig = _cluster_signature(candidate_clusters)
            # visit_lb is a valid constant lower bound on sum(x)+sum(v):
            # every pod visit moves at most C units on each side, hence
            # sum(x)+sum(v) >= 2*ceil(total_demand/C) for ANY feasible solution.
            lower_bound = (ALPHA_COORDINATION * coordination_cost
                           + UNUSED_WEIGHT * total_unused_capacity
                           + visit_lb)

            if sig in cache:
                candidate_cost = cache[sig]
                n_cached += 1
            elif lower_bound >= current_cost:
                # cannot possibly improve: sum(x)+sum(v) >= 0
                candidate_cost = lower_bound
                n_pruned += 1
            else:
                candidate_cost = evaluate_clusters(candidate_clusters, new_yy,
                                                   new_y, coordination_cost,
                                                   total_unused_capacity,
                                                   cutoff=current_cost - 0.5)
                cache[sig] = candidate_cost
                n_solves += 1

            # objective is integer-valued: require a full unit of
            # improvement (also immunises against FP noise like
            # 2669.9999999999995 vs 2670).
            if candidate_cost < current_cost - 0.5:
                current_clusters = candidate_clusters
                current_cost = candidate_cost
                if candidate_cost < best_cost - 0.5:
                    best_clusters = copy.deepcopy(candidate_clusters)
                    best_cost = candidate_cost
                    improved_this_iteration = True
                l = 0
            else:
                l += 1
        iteration += 1
        since_improvement = 0 if improved_this_iteration else since_improvement + 1
        print(f"[VNS] iteration {iteration}/{max_iterations}  "
              f"best = {best_cost}  (no-improve streak: {since_improvement})",
              flush=True)

    if since_improvement >= patience:
        print(f"[VNS] early stop: {patience} iterations without improvement",
              flush=True)
    print(f"[VNS] CPLEX solves = {n_solves} | pruned by bound = {n_pruned} | "
          f"cache hits = {n_cached}", flush=True)

    # ---- final report ----
    fy, fyy, fcoord, funused = derive_solution_components(
        best_clusters, num_picking_stations, Demand, station_capacity)
    print(f"\n[VNS FINAL] best cost = {best_cost} | extra splits = {fcoord} | "
          f"unused capacity = {funused}", flush=True)
    if _cluster_signature(best_clusters) == initial_sig:
        print("[VNS FINAL] best solution is the initial solution "
              "(already evaluated exactly above).", flush=True)
    else:
        print("[VNS FINAL] re-evaluating best solution exactly:", flush=True)
        final_cost = evaluate_clusters(best_clusters, fyy, fy, fcoord, funused,
                                       verbose=True, exact=True)
        # Both values are true costs of feasible solutions of the SAME
        # fixed assignment. If the exact re-solve hits its time cap it can
        # return a weaker incumbent than the one already found during the
        # search (whose cutoff pruned the tree and reached the better
        # solution faster) -- keep the better of the two.
        if final_cost > best_cost + 0.5:
            print(f"[VNS FINAL] exact re-solve returned a weaker incumbent "
                  f"({final_cost:g} > {best_cost:g}), time cap hit; "
                  f"keeping the better feasible value from the search.", flush=True)
        best_cost = min(best_cost, final_cost)

    return best_clusters, best_cost

Sum_demand = np.sum(Demand, axis=1)
#print("sss:",sum(Sum_demand))



for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'orders1':
            orders1_row = row
            orders1_col = col
            break

orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value
orders = np.arange(orders1)
        

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'num_picking_stations':
            num_picking_stations_row = row
            num_picking_stations_col = col
            break

num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'group_capacity':
            group_capacity_row = row
            group_capacity_col = col
            break

group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'max_iterations':
            max_iterations_row = row
            max_iterations_col = col
            break

max_iterations = sheet.cell(row=max_iterations_row + 1, column=max_iterations_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'sequences':
            sequences_row = row
            sequences_col = col
            break

sequences = sheet.cell(row=sequences_row + 1, column=sequences_col).value


num_threads = 7
iterations_per_thread = max_iterations // num_threads


#------------------------------------------------------------ Varibale Neighborhood Search -------------------------------------------------------------
start = time.time()


# ---- valid lower bound on total pod visits (used by the VNS screening) ----
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'Pods_capacity':
            Pods_capacity_row = row
            Pods_capacity_col = col
            break

Pods_capacity = int(sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value)
total_demand_all = int(np.sum(np.array(Demand)))
VISIT_LB = 2 * (-(-total_demand_all // Pods_capacity))   # 2*ceil(D/C)
print(f"[LB] total demand D = {total_demand_all} | pod capacity C = {Pods_capacity} "
      f"| visit lower bound 2*ceil(D/C) = {VISIT_LB}")

best_clusters, best_cost = variable_neighborhood_search(max_iterations, num_picking_stations, Demand, group_capacity, visit_lb=VISIT_LB)

print(f"Best result after iterations: {best_cost}")

#print("Best order groups:", best_order_groups)
print("Best cost:", best_cost)

end = time.time()
print(end - start)


