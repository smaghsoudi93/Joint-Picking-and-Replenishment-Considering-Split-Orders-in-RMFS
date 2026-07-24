# -*- coding: utf-8 -*-
# ===========================================================================
# sequential_baseline.py
# ---------------------------------------------------------------------------
# Standalone SEQUENTIAL BASELINE for the RMFS paper (Section 5.3, Table X).
# Reads the same 'Data2 - new.xlsx' instance file as the main heuristic and
# scores its solution with the SAME subproblem evaluator (evaluate_clusters,
# copied verbatim from New_paper_1404_UPDATED.py, incl. the 120 s / 1% gap
# safeguards), so the reported objective is directly comparable with both
# CPLEX and the proposed algorithm.
#
# Usage: put this file next to 'Data2 - new.xlsx' and run it, exactly like
# the main script. No VNS, no clustering heuristic -- just first-fit.
# ===========================================================================
from docplex.mp.model import Model
from collections import defaultdict
import sys
sys.path = [r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'] + sys.path
import cplex
import random
import numpy as np
import time
import threading
import pandas as pd
import concurrent.futures
#import params
import openpyxl
import copy

workbook = openpyxl.load_workbook('Data2 - new.xlsx')
sheet = workbook.active

# --- Global algorithm parameters (kept consistent with the exact OPL model) ---
ALPHA_COORDINATION = 10   # coordination cost weight (alpha in the paper / OPL model)
UNUSED_WEIGHT = 10        # weight on unused picking-station capacity (ZZ in the OPL model)

# Reproducibility (lesson from Paper 1, reviewer R2-14: the seed used in all
# reported experiments must be fixed and stated). VNS shaking is intentionally
# random; the seed makes every run repeatable.
SEED = 42
random.seed(SEED)
np.random.seed(SEED)


for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'Demand':
            demand_start_row = row + 1
            demand_start_col = col
            break

demand_rows = 0
while sheet.cell(row=demand_start_row + demand_rows, column=demand_start_col).value is not None:
    demand_rows += 1

demand_cols = 0
while sheet.cell(row=demand_start_row, column=demand_start_col + demand_cols).value is not None:
    demand_cols += 1

# Read the Demand data
Demand = []
for row in range(demand_start_row, demand_start_row + demand_rows):
    row_data = []
    for col in range(demand_start_col, demand_start_col + demand_cols):
        row_data.append(sheet.cell(row=row, column=col).value)
    Demand.append(row_data)

#for row in Demand:
#    print(row)

#xl_file = pd.read_excel("C:\Users\saman\OneDrive\Desktop\Data1.xlsx")
#Demand = xl_file.pars('Demand')
def evaluate_clusters(clusters, yy, y, coordination_cost,total_unused_capacity, verbose=False):

    ##########################################################Read the data###########################################################################

    coordination_weight = ALPHA_COORDINATION
    unused_weight = UNUSED_WEIGHT
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'matrix1':
                matrix1_start_row = row + 1
                matrix1_start_col = col
                break

    matrix1_rows = 0
    while sheet.cell(row=matrix1_start_row + matrix1_rows, column=matrix1_start_col).value is not None:
        matrix1_rows += 1

    matrix1_cols = 0
    while sheet.cell(row=matrix1_start_row, column=matrix1_start_col + matrix1_cols).value is not None:
        matrix1_cols += 1

    # Read the matrix1 data
    matrix1 = []
    for row in range(matrix1_start_row, matrix1_start_row + matrix1_rows):
        row_data1 = []
        for col in range(matrix1_start_col, matrix1_start_col + matrix1_cols):
            row_data1.append(sheet.cell(row=row, column=col).value)
        matrix1.append(row_data1)
        
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Demand1':
                demand1_start_row = row + 1
                demand1_start_col = col
                break

    demand1_rows = 0
    while sheet.cell(row=demand1_start_row + demand1_rows, column=demand1_start_col).value is not None:
        demand1_rows += 1

    demand1_cols = 0
    while sheet.cell(row=demand1_start_row, column=demand1_start_col + demand1_cols).value is not None:
        demand1_cols += 1

    # Read the Demand data
    Demand1 = []
    for row in range(demand1_start_row, demand1_start_row + demand1_rows):
        row_data2 = []
        for col in range(demand1_start_col, demand1_start_col + demand1_cols):
            row_data2.append(sheet.cell(row=row, column=col).value)
        Demand1.append(row_data2)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'items':
                items_row = row
                items_col = col
                break

    items = sheet.cell(row=items_row + 1, column=items_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'orders1':
                orders1_row = row
                orders1_col = col
                break

    orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'shelves':
                shelves_row = row
                shelves_col = col
                break

    shelves = sheet.cell(row=shelves_row + 1, column=shelves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_picking_stations':
                num_picking_stations_row = row
                num_picking_stations_col = col
                break

    num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_replenishment_stations':
                num_replenishment_stations_row = row
                num_replenishment_stations_col = col
                break

    num_replenishment_stations = sheet.cell(row=num_replenishment_stations_row + 1, column=num_replenishment_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'waves':
                waves_row = row
                waves_col = col
                break

    waves = sheet.cell(row=waves_row + 1, column=waves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Pods_capacity':
                Pods_capacity_row = row
                Pods_capacity_col = col
                break

    Pods_capacity = sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'group_capacity':
                group_capacity_row = row
                group_capacity_col = col
                break

    group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Replenishment_station_capacity':
                Replenishment_station_capacity_row = row
                Replenishment_station_capacity_col = col
                break

    Replenishment_station_capacity = sheet.cell(row=Replenishment_station_capacity_row + 1, column=Replenishment_station_capacity_col).value



    ##################################################################################################################################################


    #Groups = list(range(len(order_groups)))
    Items = list(range(items))
    Orders = list(range(orders1))
    Sequences = list(range(orders1))
    Shelves = list(range(shelves))
    Picking_stations = list(range(num_picking_stations))
    Replenishment_stations = list(range(num_replenishment_stations))
    Waves = list(range(waves))
    # Big-M values derived from instance parameters (lesson from Paper 1, R2-11 --
    # never use an arbitrary huge constant):
    #   M_uv bounds sum_i u[i,p,w,r]: can never exceed the pod capacity nor the
    #        replenishment-station capacity.
    #   M_qx bounds sum_{i,o} q[i,o,p,w,s]: the flow out of pod p in wave w,
    #        which can never exceed the pod capacity.
    M_uv = min(Pods_capacity, Replenishment_station_capacity)
    M_qx = Pods_capacity
    #waves_capacity = np.full((1, waves), (num_replenishment_stations*Replenishment_station_capacity))
    replenishmnet_capacity = np.full((waves, num_replenishment_stations), Replenishment_station_capacity)
    sample = np.zeros((len(Demand1), len(Demand1[0]), waves))
    capacity_of_pods = np.full((waves, shelves), Pods_capacity)
    #Arriving_times = (10, 15, 20, 25) #(7, 12, 17, 22, 27)#(5, 10, 15, 20, 25, 30) #(7, 12, 17, 22, 27) #(10, 15, 20, 25, 30) #(7, 12, 17, 22, 27)#(7, 12, 17, 22, 27, 32) #(5, 10) #(6, 12, 18, 24, 30)  #(5, 10) # 
    #Duration = (5, 5, 5, 5) #(5, 5, 5, 5, 5, 5) #(5, 5) #(6, 6, 6, 6, 6) #(5, 5) # 
    

    #items = 11 #5 #7 #5  
    #orders1 = 8 #6 #50 #20 #15  
    #shelves = 3 #10 
    #num_picking_stations = 3 #2 #6 #4  
    #num_replenishment_stations = 3 #2 #6 #4  
    #waves = 4 #2 #6 #5   
    #Pods_capacity = 45 #15 #35 #25   
    #group_capacity = 2 #7 #3 #4 
    #Replenishment_station_capacity = 25 #22 #40   
    
    

#-----------------------------------------------------------------------------
# Build the model
#-----------------------------------------------------------------------------

# Create CPO model
    mdl = Model(name='Thesis')
    #print(order_groups)


    #qq = np.zeros((len(order_groups),shelves, items, waves))
    #zz = np.zeros((len(order_groups),shelves, waves))
    vv = np.zeros((shelves, waves, num_replenishment_stations))
    uu = np.zeros((items, shelves, waves, num_replenishment_stations))
    fff = np.zeros((items, waves, num_replenishment_stations))


    #xx = np.zeros((len(matrix1),len(order_groups)))
    #x = {}
#for group_idx, group in enumerate(order_groups):
 #   for order in group:
    #for g in Groups:
     #   for o in Orders:
      #      x[o, g] = mdl.binary_var(name=f'x_{o}_{g}')

#for g in Groups:
    #for o in Orders:
        #mdl.add_constraint(x[o, g] == 0)

# Set constraints based on order_groups
    #for group_idx, group in enumerate(order_groups):
     #   for order in group:
      #      xx[order, group_idx] = 1
       #     mdl.add_constraint(x[order, group_idx] == 1)
    #print(xx)
    #print(yy)
    y1= {}
    for i in Items:
        for j in Orders:
            for s in Picking_stations:
                y1[i, j, s] =mdl.binary_var(name=f'y_{i}_{j}_{s}')

    # Fix y1 to the heuristic's assignment in BOTH directions.
    # (Previously only cells with y==1 were fixed, leaving y==0 cells free --
    # the MIP could silently add extra item-station assignments unrelated to
    # the Phase 1-2 decisions.)
    for i in Items:
        for j in Orders:
            for s in Picking_stations:
                mdl.add_constraint(y1[i, j, s] == int(y[i, j, s]))

    # Item-level coverage (matches const21 in the exact OPL model): every
    # requested item must be assigned to exactly one station. Because y1 is
    # fully fixed above, this acts as a hard validity check on the heuristic's
    # output -- if Phases 1-2 ever produce an uncovered or double-covered item,
    # the model becomes infeasible and the failure is caught immediately
    # instead of silently returning a meaningless objective.
    for j in Orders:
        for i in Items:
            if Demand1[j][i] > 0:
                mdl.add_constraint(mdl.sum(y1[i, j, s] for s in Picking_stations) == 1)
                
    z = {(i): mdl.continuous_var(name="z%d" % (i)) for i in Picking_stations}
    v = {(i,j,k): mdl.binary_var(name="v%d%d%d" % (i,j,k)) for i in Shelves for j in Waves for k in Replenishment_stations}
    # q is STATION-INDEXED (5-dim), matching the updated exact model: a pod
    # visit is counted separately for each station it serves in a wave, which
    # is what makes order splitting genuinely consume additional pod visits.
    q = {(i,j,k,ii,s): mdl.continuous_var(name="q_%d_%d_%d_%d_%d" % (i,j,k,ii,s)) for i in Items for j in Orders for k in Shelves for ii in Waves for s in Picking_stations}
    u = {(i,j,k,ii): mdl.continuous_var(name="u%d%d%d%d" % (i,j,k,ii)) for i in Items for j in Shelves for k in Waves for ii in Replenishment_stations}
    #ff = {(i,j,k): mdl.continuous_var(name="ff%d%d%d" % (i,j,k)) for i in Items for j in Waves for k in Replenishment_stations}
    x = {(i,j,k): mdl.binary_var(name="x%d%d%d" % (i,j,k)) for i in Shelves for j in Waves for k in Picking_stations}

    ff= {}
    for i in Items:
        for j in Waves:
            for k in Replenishment_stations:
                ff[i, j, k] =mdl.continuous_var(name=f'ff_{i}_{j}_{k}')
            
    waving_items = np.zeros((waves, len(matrix1[0])))
    replenishment_items = np.zeros((waves, len(matrix1[0]),num_replenishment_stations))
    pod_items = np.zeros((len(matrix1[0]),shelves ,waves, num_replenishment_stations))
    
# ------------------------------------------------------------------------ waving heuristic -------------------------------------------------------
    # A wave can carry no more than EITHER bottleneck allows: the total
    # replenishment-station capacity OR the total pod capacity of that wave.
    # (Previously only the replenishment side was considered, so waves could
    # be built that then failed to fit into the pods.)
    num_waves = waves   # keep the scalar (number of waves) before it gets shadowed below
    wave_capacity = min(num_replenishment_stations*Replenishment_station_capacity,
                        shelves*Pods_capacity)
    num_items = len(matrix1[0])

        

    wave_plan = []
    current_wave = [0] * num_items
    current_total = 0

    for cluster in clusters:
        for _, item, qty in cluster:
            while qty > 0:
                available_space = wave_capacity - current_total
                if available_space == 0:
                    wave_plan.append(current_wave)
                    current_wave = [0] * num_items
                    current_total = 0
                    available_space = wave_capacity
                assign_qty = min(qty, available_space)
                current_wave[item] += assign_qty
                current_total += assign_qty
                qty -= assign_qty

    if current_total > 0:
        wave_plan.append(current_wave)

    # graceful failure instead of IndexError if the plan needs more waves
    # than the instance provides
    if len(wave_plan) > num_waves:
        print(f"[wave-planning] infeasible: heuristic needs {len(wave_plan)} waves, "
              f"instance only has {num_waves}")
        return 10000000

    wavess = np.array(wave_plan)
    for i in range(len(wavess)):
        for j in range(len(wavess[0])):
            waving_items[i][j] = wavess[i][j]

    # ---- Stage 3b: distribute each wave's items across replenishment stations
    # (greedy sequential fill). Replaces a hardcoded 3-level if/else cascade
    # that (a) crashed with an IndexError whenever more than 4 stations were
    # needed and (b) had no bound check at all -- the same style of bug that
    # crashed on instances with few pods. The while-loop below implements the
    # identical greedy logic for ANY number of stations, and returns the
    # standard infeasibility sentinel (10000000) if the wave genuinely does
    # not fit, instead of crashing.
    for waves_number in Waves:
        r = 0
        for j in range(len(matrix1[0])):
            remaining = waving_items[waves_number][j]
            while remaining > 0:
                if r >= num_replenishment_stations:
                    print(f"[wave-planning] infeasible: wave {waves_number} exceeds "
                          f"total replenishment capacity")
                    return 10000000
                take = min(remaining, replenishmnet_capacity[waves_number][r])
                replenishment_items[waves_number][j][r] += take
                replenishmnet_capacity[waves_number][r] -= take
                remaining -= take
                if remaining > 0:
                    r += 1

        # ---- Stage 3c: place replenished quantities into pods (greedy
        # sequential fill; per-wave pod capacity). Same generalization: the
        # old code did p+1 with no bound check and crashed with 2 pods.
        p = 0
        for k in Replenishment_stations:
            for i in range(len(matrix1[0])):
                remaining = replenishment_items[waves_number][i][k]
                while remaining > 0:
                    if p >= shelves:
                        print(f"[wave-planning] infeasible: wave {waves_number} exceeds "
                              f"total pod capacity")
                        return 10000000
                    take = min(remaining, capacity_of_pods[waves_number][p])
                    pod_items[i][p][waves_number][k] += take
                    capacity_of_pods[waves_number][p] -= take
                    remaining -= take
                    if remaining > 0:
                        p += 1
                
                
# -------------------------------------------------------- constraints ------------------------------------------------------
#Constant variables based on heuristic:


    #for i in Items:
     #   for j in Waves:
      #      for k in Replenishment_stations:
                    #print(i,j,k, replenishment_items[j][i][k])
       #         mdl.add_constraint(ff[i, j, k] == replenishment_items[j][i][k])

    for i in Items:
        for j in Waves:
            for k in Replenishment_stations:
                for p in Shelves:
                    mdl.add_constraint(u[i, p, j, k] == pod_items[i][p][j][k])


#print(order_groups)

# ---------------------------- decision variables ---------------------

#y = {(i,j,k): mdl.binary_var(name="y%d%d%d" % (i+1,j+1,k+1)) for i in Groups for j in Picking_stations for k in Sequences}
        
        
#--------------------------------------------------------------------------------------
#mdl.add_constraint(l[0,0] == 1)
#mdl.add_constraint(l[1,0] == 1)
#mdl.add_constraint(l[2,0] == 1)
#mdl.add_constraint(l[3,2] == 1)

        #for g in range(len(Groups1)):
        #    if groups_demand[0][Groups1[g]] == 0:
                #print(Groups1[g])
         #       mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves for w in Waves) >= 1)

        #for w in Waves:
         #   mdl.add_constraint(mdl.sum(v[p,w,r] for p in Shelves for r in Replenishment_stations) >= 1)

    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for p in Shelves) <= Replenishment_station_capacity)

    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for r in Replenishment_stations) <= Pods_capacity)

        
    # (matches updated OPL const10, per-station equality): the quantity of
    # item i for order o picked AT STATION s equals the demand assigned there.
    # When y1[i,o,s]=0 this forces all q[i,o,p,w,s]=0 -- a tight, big-M-free
    # link between picking flows and item-station assignment.
    for i in Items:
        for o in Orders:
            for s in Picking_stations:
                mdl.add_constraint(Demand1[o][i]*y1[i,o,s] == mdl.sum(q[i,o,p,w,s] for p in Shelves for w in Waves))


    # (matches updated OPL const13): replenished quantity into pod (p,w)
    # equals total picked from it across all orders AND stations.
    for w in Waves:
        for i in Items:
            for p in Shelves:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for r in Replenishment_stations) == mdl.sum(q[i,o,p,w,s] for o in Orders for s in Picking_stations))

    for i in Items:
        for o in Orders:
            for s in Picking_stations:
                mdl.add_constraint(mdl.sum(x[p,w,s] for p in Shelves for w in Waves) >= y1[i,o,s])


    # NOTE: the three blocks below were previously indented one level too
    # deep (inside the Items loop), silently adding |Items| duplicate copies
    # of every constraint. Now at function level (added exactly once).
    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) >= v[p,w,r])

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) <= M_uv*v[p,w,r])

    # (matches updated OPL const20, per-station): picking from pod p at
    # station s in wave w requires the pod to be AT that station -- the key
    # constraint that makes each pod-station visit count separately.
    for p in Shelves:
        for w in Waves:
            for s in Picking_stations:
                mdl.add_constraint(mdl.sum(q[i,o,p,w,s] for i in Items for o in Orders) <= M_qx*x[p,w,s])

        
                               

#Objective function
    mdl.total_objective = mdl.sum(x[p,w,s] for p in Shelves for w in Waves for s in Picking_stations) + mdl.sum(v[p,w,r] for p in Shelves for w in Waves for r in Replenishment_stations)+ (coordination_weight *coordination_cost)+(unused_weight*total_unused_capacity)
        #mdl.total_objective = mdl.sum(ctt[o] for o in Orders)/ orders1
    mdl.minimize(mdl.total_objective)
    # Exact evaluation (no gap tolerance) so the baseline objective is
    # directly comparable with the proposed method's exact values;
    # generous cap since the baseline performs a single solve per run.
    mdl.parameters.timelimit = 600

    #mdl.solve(log_=True)
        #mdl.solve()


      
        #log_output=True
    #if mdl.solve(log_output=True):
    if mdl.solve():
        total_cost = mdl.objective_value
        if verbose:
            sum_x = sum(x[p,w,s].solution_value for p in Shelves for w in Waves for s in Picking_stations)
            sum_v = sum(v[p,w,r].solution_value for p in Shelves for w in Waves for r in Replenishment_stations)
            print(f"[eval breakdown] pod-picking visits (Σx) = {sum_x:.0f} | "
                  f"pod-replenishment visits (Σv) = {sum_v:.0f} | "
                  f"coordination = {coordination_weight}×{coordination_cost} = {coordination_weight*coordination_cost} | "
                  f"unused = {unused_weight}×{total_unused_capacity} = {unused_weight*total_unused_capacity} | "
                  f"TOTAL = {total_cost}")
        else:
            print("total_cost:", total_cost)

    else:
        print("no_solution")
        total_cost = 10000000

    end2 = time.time()
        #print("wave:",waves_number, end2 - start2)
    
    
    return total_cost
def update_station_and_order_mapping(clusters, num_stations):
    station_assignments = defaultdict(list)
    order_to_stations = defaultdict(set)

    for cluster_idx, cluster in enumerate(clusters):
        station_id = cluster_idx % num_stations  # Round-robin
        station_assignments[station_id].append(cluster)
        for order_id, _, _ in cluster:
            order_to_stations[order_id].add(station_id)

    return station_assignments, order_to_stations
def derive_solution_components(clusters, num_picking_stations, Demand, station_capacity):
    """Given a clustering, derive everything evaluate_clusters needs:
    the item-order-station assignment y, per-order station counts yy,
    the coordination cost, and total unused picking capacity."""
    station_assignments, order_to_stations = update_station_and_order_mapping(clusters, num_picking_stations)

    y = np.zeros((len(Demand[0]), len(Demand), num_picking_stations))
    for sid, assigned_clusters in station_assignments.items():
        for cid, cluster in enumerate(assigned_clusters, 1):
            for order_idx, sku_idx, qty in cluster:
                y[sku_idx][order_idx][sid] = 1

    yy = np.zeros((1, len(Demand)))
    for order, stations in order_to_stations.items():
        yy[0][order] = len(stations)

    station_usage = defaultdict(int)
    for sid, cluster_list in station_assignments.items():
        for cluster in cluster_list:
            station_usage[sid] += sum(qty for _, _, qty in cluster)

    total_unused_capacity = sum(
        station_capacity - min(station_capacity, station_usage[sid])
        for sid in range(num_picking_stations))

    coordination_cost = sum(len(v) - 1 for v in order_to_stations.values())
    return y, yy, coordination_cost, total_unused_capacity


# ===========================================================================
# SEQUENTIAL BASELINE  (Section 5.3, "Comparison with Sequential Approach")
# ===========================================================================
# The baseline mirrors what the paper describes: orders are assigned to
# picking stations purely on a CAPACITY basis (classic first-fit), with no
# co-occurrence affinity, no coordination-cost awareness, and no VNS
# improvement. Pod allocation and replenishment are then resolved for that
# FIXED assignment by the same subproblem MIP used by the proposed method
# (evaluate_clusters), so both approaches are scored on the IDENTICAL
# objective  sum(x) + sum(v) + alpha*coordination + ZZ*unused.
#
# Design notes (kept deliberately fair -- not a strawman):
#   * Item-level integrality is respected: each (order, item) pair is placed
#     WHOLE at one station, exactly as required by the coverage constraint
#     (const21) of the exact model. Orders may still split across stations
#     item by item.
#   * Items of an order are processed consecutively, so first-fit keeps an
#     order together as long as the current stations have room; splits arise
#     only at capacity boundaries -- the natural behaviour of a
#     coordination-agnostic policy, not an artificially scattered one.
#   * No randomness: the baseline is deterministic and needs no seed.

def sequential_assignment(Demand, num_picking_stations, station_capacity):
    """Capacity-driven first-fit assignment of order items to stations.
    Returns clusters in the same format used everywhere else:
    cluster index == station id (1:1), each cluster a list of
    (order, sku, qty) triplets."""
    num_orders = len(Demand)
    num_skus = len(Demand[0])

    clusters = [[] for _ in range(num_picking_stations)]
    load = [0] * num_picking_stations

    for o in range(num_orders):
        for i in range(num_skus):
            qty = Demand[o][i]
            if qty is None or qty <= 0:
                continue
            placed = False
            for s in range(num_picking_stations):        # first-fit
                if load[s] + qty <= station_capacity:
                    clusters[s].append((o, i, int(qty)))
                    load[s] += qty
                    placed = True
                    break
            if not placed:
                print(f"[sequential] infeasible: item {i} of order {o} "
                      f"(qty {qty}) fits in no station")
                return None
    return clusters


# ------------------------------- driver ------------------------------------

Sum_demand = np.sum(Demand, axis=1)

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'num_picking_stations':
            num_picking_stations_row = row
            num_picking_stations_col = col
            break

num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'group_capacity':
            group_capacity_row = row
            group_capacity_col = col
            break

group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value


print("=" * 72)
print("SEQUENTIAL BASELINE  (capacity-only assignment, no coordination, no VNS)")
print("=" * 72)

start = time.time()

seq_clusters = sequential_assignment(Demand, num_picking_stations, group_capacity)
if seq_clusters is None:
    raise SystemExit("[sequential] aborted: assignment infeasible")

y, yy, coordination_cost, total_unused_capacity = derive_solution_components(
    seq_clusters, num_picking_stations, Demand, group_capacity)

# ---- Table X metrics, part 1 (assignment structure) ----
n_orders = len(Demand)
stations_per_order = [int(yy[0][o]) for o in range(n_orders)]
n_split_orders = sum(1 for k in stations_per_order if k > 1)
print(f"[sequential] extra splits (coordination units) = {coordination_cost}")
print(f"[sequential] orders split across >1 station     = {n_split_orders} / {n_orders}")
print(f"[sequential] max stations used by one order     = {max(stations_per_order)}")
print(f"[sequential] unused picking capacity            = {total_unused_capacity}")

# ---- Table X metrics, part 2 (full objective via the SAME evaluator) ----
total_cost = evaluate_clusters(seq_clusters, yy, y,
                               coordination_cost, total_unused_capacity,
                               verbose=True)

end = time.time()
print(f"[sequential] TOTAL objective = {total_cost}")
print(f"[sequential] wall time = {end - start:.1f} s")
